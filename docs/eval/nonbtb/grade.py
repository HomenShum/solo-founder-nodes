#!/usr/bin/env python3
"""Deterministic grader for the non-BTB generalization slice.

Mirrors the BTB scorecard dimensions WITHOUT an LLM verifier, so it is cheap,
reproducible, and cannot be gamed by per-task answer-writers:
  - correctness   : produced value within tolerance of the golden value
  - formula       : the value is a live spreadsheet formula, not a hardcoded literal
  - cited         : the value carries a citation that resolves to a real source file
  - fabrication   : any produced key outside the allowed set, or any citation to a
                    non-existent source, is penalized

Contract: the agent writes, into <output_dir>:
  - outputs.json : { metric_key: { "value": <num>, "formula": "=...", "cite": {"file": "...", "locator": "..."} }, ... }
  - (optional) the human deliverable named by rubric["deliverable"] (xlsx checked for "=" cells)

Usage:
  python grade.py <task_dir> <output_dir>
  python grade.py --all <eval_root> <outputs_root>   # outputs_root/<task>/outputs.json
"""
import json, os, sys

def load_json(path):
    with open(path, encoding="utf-8-sig") as f:
        return json.load(f)

def grade_task(task_dir, output_dir):
    rubric = load_json(os.path.join(task_dir, "rubric.json"))
    allowed = set(rubric["allowed_keys"])
    expected = rubric["expected"]
    sources = set(rubric.get("sources", []))
    need_formula = rubric.get("formula_required", False)
    need_cite = rubric.get("citations_required", False)

    out_path = os.path.join(output_dir, "outputs.json")
    if not os.path.exists(out_path):
        return {"task": rubric["task"], "score": 0.0, "note": "no outputs.json produced",
                "correct": 0, "formula": 0, "cited": 0, "fabrication": 0, "n": len(expected)}
    outputs = load_json(out_path)

    n = len(expected)
    correct = formula = cited = 0
    fabrication = 0
    detail = {}
    for key, spec in expected.items():
        rec = outputs.get(key)
        if not isinstance(rec, dict):
            detail[key] = "missing"
            continue
        ok_val = isinstance(rec.get("value"), (int, float)) and abs(rec["value"] - spec["value"]) <= spec.get("tol", 0.0)
        ok_formula = (not need_formula) or (isinstance(rec.get("formula"), str) and rec["formula"].lstrip().startswith("="))
        cite = rec.get("cite") or {}
        ok_cite = (not need_cite) or (cite.get("file") in sources)
        correct += int(ok_val); formula += int(ok_formula); cited += int(ok_cite)
        detail[key] = {"value_ok": ok_val, "formula_ok": ok_formula, "cited_ok": ok_cite}

    # fabrication: keys outside the allowed set, or citations to non-source files
    for key, rec in outputs.items():
        if key not in allowed:
            fabrication += 1
        elif isinstance(rec, dict):
            f = (rec.get("cite") or {}).get("file")
            if f is not None and f not in sources:
                fabrication += 1

    # optional: verify deliverable xlsx actually contains formulas (defense vs json-only claims)
    deliverable_formula_ok = None
    deliv = rubric.get("deliverable", "")
    dpath = os.path.join(output_dir, deliv) if deliv else ""
    if need_formula and deliv.endswith(".xlsx") and os.path.exists(dpath):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(dpath)
            has_formula = any(
                isinstance(c.value, str) and c.value.lstrip().startswith("=")
                for ws in wb.worksheets for row in ws.iter_rows() for c in row
            )
            deliverable_formula_ok = bool(has_formula)
        except Exception as e:
            deliverable_formula_ok = f"error: {e}"

    dims = correct + (formula if need_formula else 0) + (cited if need_cite else 0)
    denom = n * (1 + int(need_formula) + int(need_cite))
    raw = dims / denom if denom else 0.0
    score = max(0.0, raw - 0.1 * fabrication)  # each fabrication costs 10%
    return {
        "task": rubric["task"], "n": n, "correct": correct,
        "formula": formula if need_formula else None,
        "cited": cited if need_cite else None,
        "fabrication": fabrication,
        "deliverable_formula_ok": deliverable_formula_ok,
        "score": round(score, 3), "detail": detail,
    }

def main(argv):
    if len(argv) >= 1 and argv[0] == "--all":
        eval_root, outputs_root = argv[1], argv[2]
        tasks = sorted(d for d in os.listdir(eval_root)
                       if os.path.exists(os.path.join(eval_root, d, "rubric.json")))
        results = [grade_task(os.path.join(eval_root, t), os.path.join(outputs_root, t)) for t in tasks]
        mean = sum(r["score"] for r in results) / len(results) if results else 0.0
        for r in results:
            print(f"{r['task']:28} score={r['score']:.3f}  correct={r['correct']}/{r['n']} "
                  f"formula={r['formula']} cited={r['cited']} fab={r['fabrication']}")
        print(f"\nNON-BTB SLICE MEAN: {mean:.3f}  (n_tasks={len(results)})")
        return 0
    task_dir, output_dir = argv[0], argv[1]
    print(json.dumps(grade_task(task_dir, output_dir), indent=2))
    return 0

if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
